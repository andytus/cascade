from django.contrib.sites.models import Site
from django.core.files import File
from django.contrib.sites.managers import CurrentSiteManager
from django.contrib.contenttypes.models import ContentType
from django.conf import settings
from django.core.urlresolvers import reverse
from django.core.exceptions import ValidationError
from django.utils.safestring import mark_safe
from django.utils import timezone
from django.db import models
from django.db.models import Avg, Min, Max, Count, Sum
from django.db.models.signals import post_save
from cascade.apps.report_builder.unique_slugify import unique_slugify
from cascade.apps.report_builder.utils import get_model_from_path_string, report_to_list
from dateutil import parser
import os

def save_report_to(instance, filename):
    """
     Gets correct path for an uploaded file.
     Will upload to the media folder../reports/<site__name>, report name, file_name
    """

    return os.path.join('reports', instance.site.name, instance.report.name, filename)


AUTH_USER_MODEL = getattr(settings, 'AUTH_USER_MODEL', 'auth.User')


class Report(models.Model):
    """ A saved report with queryset and descriptive fields
    """
    def _get_model_manager(self):
        """
        Get default manager from settings else use objects
        """
        model_manager = 'objects'  #sets the default to objects
        if getattr(settings, 'REPORT_BUILDER_MODEL_MANAGER', False):
            model_manager = settings.REPORT_BUILDER_MODEL_MANAGER
        return model_manager

    def _get_allowed_models():
        models = ContentType.objects.all()
        if getattr(settings, 'REPORT_BUILDER_INCLUDE', False):
            models = models.filter(name__in=settings.REPORT_BUILDER_INCLUDE)
        if getattr(settings, 'REPORT_BUILDER_EXCLUDE', False):
            models = models.exclude(name__in=settings.REPORT_BUILDER_EXCLUDE)
        return models
    
    name = models.CharField(max_length=255)
    slug = models.SlugField(verbose_name="Short Name")
    description = models.TextField(blank=True)
    root_model = models.ForeignKey(ContentType, limit_choices_to={'pk__in':_get_allowed_models})
    created = models.DateField(auto_now_add=True)
    modified = models.DateField(auto_now=True)
    user_created = models.ForeignKey(AUTH_USER_MODEL, editable=False, blank=True, null=True)
    user_modified = models.ForeignKey(AUTH_USER_MODEL, editable=False, blank=True, null=True, related_name="report_modified_set")
    distinct = models.BooleanField(default=False)
    starred = models.ManyToManyField(AUTH_USER_MODEL, blank=True,
                                     help_text="These users have starred this report for easy reference.",
                                     related_name="report_starred_set")
    site = models.ManyToManyField(Site)
    objects = models.Manager()
    on_site = CurrentSiteManager()

    
    def save(self, *args, **kwargs):
        if not self.id:
            unique_slugify(self, self.name)
        super(Report, self).save(*args, **kwargs)


    def add_aggregates(self, queryset):
        for display_field in self.displayfield_set.filter(aggregate__isnull=False):
            if display_field.aggregate == "Avg":
                queryset = queryset.annotate(Avg(display_field.path + display_field.field))
            elif display_field.aggregate == "Max":
                queryset = queryset.annotate(Max(display_field.path + display_field.field))
            elif display_field.aggregate == "Min":
                queryset = queryset.annotate(Min(display_field.path + display_field.field))
            elif display_field.aggregate == "Count":
                queryset = queryset.annotate(Count(display_field.path + display_field.field))
            elif display_field.aggregate == "Sum":
                queryset = queryset.annotate(Sum(display_field.path + display_field.field))
        return queryset

    
    def get_query(self, site=None):
        report = self
        model_class = report.root_model.model_class()
        message= ""
        if site:
            objects = model_class.objects.filter(site=site)
        else:
            if getattr(model_class, 'report_builder_model_manager', False):
                objects = getattr(model_class, 'report_builder_model_manager').all()
            else:
                manager = report._get_model_manager()
                objects = getattr(model_class, manager).all()  #changed from model_class.objects.all()

        # Filters
        # NOTE: group all the filters together into one in order to avoid 
        # unnecessary joins
        filters = {}
        excludes = {}
        for filter_field in report.filterfield_set.all():
            try:
                # exclude properties from standard ORM filtering 
                if '[property]' in filter_field.field_verbose:
                    continue
                if '[custom' in filter_field.field_verbose:
                    continue

                filter_string = str(filter_field.path + filter_field.field)
                
                if filter_field.filter_type:
                    filter_string += '__' + filter_field.filter_type
                
                # Check for special types such as isnull
                if filter_field.filter_type == "isnull" and filter_field.filter_value == "0":
                    filter_ = {filter_string: False}
                elif filter_field.filter_type == "in":
                    filter_ = {filter_string: filter_field.filter_value.split(',')}
                else:
                    # All filter values are stored as strings, but may need to be converted
                    if '[Date' in filter_field.field_verbose:
                        filter_value = parser.parse(filter_field.filter_value)
                        if settings.USE_TZ:
                            filter_value = timezone.make_aware(
                                filter_value,
                                timezone.get_current_timezone()
                            )
                        if filter_field.filter_type == 'range':
                            filter_value = [filter_value, parser.parse(filter_field.filter_value2)]
                            if settings.USE_TZ:
                                filter_value[1] = timezone.make_aware(
                                    filter_value[1],
                                    timezone.get_current_timezone()
                                )
                    else:
                        filter_value = filter_field.filter_value
                        if filter_field.filter_type == 'range':
                            filter_value = [filter_value, filter_field.filter_value2]
                    filter_ = {filter_string: filter_value}

                if not filter_field.exclude:
                    filters.update(filter_) 
                else:
                    excludes.update(filter_) 

            except Exception:
                import sys
                e = sys.exc_info()[1]
                message += "Filter Error on %s. If you are using the report builder then " % filter_field.field_verbose
                message += "you found a bug! "
                message += "If you made this in admin, then you probably did something wrong."

        if filters:
            objects = objects.filter(**filters)
        if excludes:
            objects = objects.exclude(**excludes)

        # Aggregates
        objects = self.add_aggregates(objects) 

        # Distinct
        if report.distinct:
            objects = objects.distinct()

        return objects, message
    
    @models.permalink
    def get_absolute_url(self):
        return ("report_update_view", [str(self.id)])

    def generate_report_xlwt(self, user, site):
        import xlwt
        import cStringIO as StringIO
        import re

        report = self
        wb = xlwt.Workbook()
        ws = wb.add_sheet(report.name)
        filename = re.sub(r'\W+', '', report.name) + "_" + site.name+ '.xlsx'

        i = 0
        for field in report.displayfield_set.all():
            ws.write(0, i, field.name, xlwt.easyxf("font: bold on"))
            i += 1
        objects_list, message = report_to_list(report, user, site, queryset=None)

        r = 1
        for row in objects_list:
            c = 0
            for value in row:
                ws.write(r, c, value)
                c += 1
            r += 1

        myfile = StringIO.StringIO()
        wb.save(myfile)

        # a unique file will be related to this report, file_extension, and site
        report_file, created = ReportFiles.objects.get_or_create(report=self, file_extension='xlsx', site=site)
        if created:
            report_file.save()
        else:
            report_file.file_path.delete()

        report_file.file_path.save(filename, File(myfile))
        report_file.update_in_progress = 'No'
        report_file.save()

    def generate_report(self, user, site):
        from django.db import connection
        connection.close()
        import cStringIO as StringIO
        from openpyxl.workbook import Workbook
        from openpyxl.writer.excel import save_virtual_workbook
        from openpyxl.cell import get_column_letter
        import re

        report = self
        wb = Workbook()
        ws = wb.worksheets[0]
        ws.title = re.sub(r'\W+', '', report.name)[:30]
        filename = re.sub(r'\W+', '', report.name) + "_" + site.name + '.xlsx'

        i = 0
        for field in report.displayfield_set.all():
            cell = ws.cell(row=0, column=i)
            cell.value = field.name
            cell.style.font.bold = True
            ws.column_dimensions[get_column_letter(i + 1)].width = field.width
            i += 1

        objects_list, message = report_to_list(report, user, site, queryset=None)
        for row in objects_list:
            try:
                ws.append(row)
            except ValueError as e:
                ws.append([e.message])
            except:
                ws.append(['Unknown Error'])

        myfile = StringIO.StringIO()
        myfile.write(save_virtual_workbook(wb))

        # a unique file will be related to this report, file_extension, and site
        report_file, created = ReportFiles.objects.get_or_create(report=self, file_extension='xlsx', site=site)
        if created:
            report_file.save()
        else:
            report_file.file_path.delete()

        report_file.file_path.save(filename, File(myfile))
        report_file.update_in_progress = 'No'
        report_file.save()

    def generate_report_sites(self, user):
        for s in self.site.all():
            self.generate_report(user, s)

    def check_report_display_field_positions(self):
        """ After report is saved, make sure positions are sane
        """
        for i, display_field in enumerate(self.displayfield_set.all()):
            if display_field.position != i+1:
                display_field.position = i+1
                display_field.save()

    def __unicode__(self):
        return "%s : %s" % (self.name, self.root_model)

class ReportFiles(models.Model):
    FILE_TYPE = (('xlsx', 'xlsx'), ('csv', 'csv'))
    """
    Holds path and meta data for all report files on all sites
    """
    report = models.ForeignKey(Report, related_name='report_file')
    file_extension = models.CharField(choices=FILE_TYPE, default="xlsx", max_length=8)
    file_path = models.FileField(upload_to=save_report_to, null=True)
    last_generated = models.DateTimeField(auto_now=True)
    last_generated_by = models.ForeignKey(AUTH_USER_MODEL, blank=True, null=True)
    update_in_progress = models.CharField(choices=(('Yes', 'Yes'), ('No', 'No')), default='No', max_length=4)
    site = models.ForeignKey(Site)
    objects = models.Manager()
    on_site = CurrentSiteManager()

    class Meta:
        verbose_name_plural = "Report Files"

class Format(models.Model):
    """ A specifies a Python string format for e.g. `DisplayField`s. 
    """
    name = models.CharField(max_length=50, blank=True, default='')
    string = models.CharField(max_length=300, blank=True, default='', help_text='Python string format. Ex ${} would place a $ in front of the result.')

    def __unicode__(self):
        return self.name
    

class DisplayField(models.Model):
    """ A display field to show in a report. Always belongs to a Report
    """
    report = models.ForeignKey(Report)
    path = models.CharField(max_length=2000, blank=True)
    path_verbose = models.CharField(max_length=2000, blank=True)
    field = models.CharField(max_length=2000)
    field_verbose = models.CharField(max_length=2000)
    name = models.CharField(max_length=2000)
    sort = models.IntegerField(blank=True, null=True)
    sort_reverse = models.BooleanField(verbose_name="Reverse", default=False)
    width = models.IntegerField(default=15)
    aggregate = models.CharField(
        max_length=5,
        choices = (
            ('Sum','Sum'),
            ('Count','Count'),
            ('Avg','Avg'),
            ('Max','Max'),
            ('Min','Min'),
        ),
        blank = True
    )
    position = models.PositiveSmallIntegerField(blank = True, null = True)
    total = models.BooleanField(default=False)
    group = models.BooleanField(default=False)
    display_format = models.ForeignKey(Format, blank=True, null=True)

    class Meta:
        ordering = ['position']
    
    def get_choices(self, model, field_name):
        try:
            model_field = model._meta.get_field_by_name(field_name)[0]
        except:
            model_field = None
        if model_field and model_field.choices:
            return model_field.choices

    @property
    def choices_dict(self):
        choices = self.choices
        choices_dict = {}
        if choices:
            for choice in choices:
                choices_dict.update({choice[0]: choice[1]})
        return choices_dict

    @property
    def choices(self):
        if self.pk:
            model = get_model_from_path_string(self.report.root_model.model_class(), self.path)
            return self.get_choices(model, self.field)

    def __unicode__(self):
        return self.name
        
class FilterField(models.Model):
    """ A display field to show in a report. Always belongs to a Report
    """
    report = models.ForeignKey(Report)
    path = models.CharField(max_length=2000, blank=True)
    path_verbose = models.CharField(max_length=2000, blank=True)
    field = models.CharField(max_length=2000)
    field_verbose = models.CharField(max_length=2000)
    filter_type = models.CharField(
        max_length=20,
        choices = (
            ('exact','Equals'),
            ('iexact','Equals (case-insensitive)'),
            ('contains','Contains'),
            ('icontains','Contains (case-insensitive)'),
            ('in','in (comma seperated 1,2,3)'),
            ('gt','Greater than'),
            ('gte','Greater than equals'),
            ('lt','Less than'),
            ('lte','Less than equals'),
            ('startswith','Starts with'),
            ('istartswith','Starts with (case-insensitive)'),
            ('endswith','Ends with'),
            ('iendswith','Ends with  (case-insensitive)'),
            ('range','range'),
            ('week_day','Week day'),
            ('isnull','Is null'),
            ('regex','Regular Expression'),
            ('iregex','Reg. Exp. (case-insensitive)'),
        ),
        blank=True,
        default = 'icontains',
    )
    filter_value = models.CharField(max_length=2000)
    filter_value2 = models.CharField(max_length=2000, blank=True)
    exclude = models.BooleanField(default=False)
    position = models.PositiveSmallIntegerField(blank = True, null = True)

    class Meta:
        ordering = ['position']
    
    def clean(self):
        if self.filter_type == "range":
            if self.filter_value2 in [None, ""]:
                raise ValidationError('Range filters must have two values')
        return super(FilterField, self).clean()


    def get_choices(self, model, field_name):
        try:
            model_field = model._meta.get_field_by_name(field_name)[0]
        except:
            model_field = None
        if model_field and model_field.choices:
            return model_field.choices

    @property
    def choices(self):
        if self.pk:
            model = get_model_from_path_string(self.report.root_model.model_class(), self.path)
            return self.get_choices(model, self.field)

    def __unicode__(self):
        return self.field